"""GSA Corp "Certified Appraisal Roll" xlsx loader — the free roll export GSA
Corp CADs post as a zipped .xlsx (sheet "SearchResults", ~60 cols, ONE row per
account = the primary improvement). Verified: Ector (ectorcad.org/home/downloads).
Same vendor as Johnson/Smith; reusable for any GSA county that posts the xlsx.

Columns used: ACCOUNT_NUMBER (dotted) + GIS_IDENTIFICATION_NUM (dashed) for the
join auto-detect; DATE_OF_SALE, EXEMPTION (clean codes: HS/TOT/OV65/DV...),
IMPROVEMENT_TYPE (-> improvements[] + pool/garage/shed tags), IMPROVEMENT_AREA
(-> sqft), YEAR_BUILT. (No beds/baths in this export; EXTRA_FEATURE_VALUE is a $
amount, not a marker, so skipped.)

JOIN AUTO-DETECT (Fort Bend lesson): test account/geoid x source_property_id/apn,
pick the highest, require >=30% or abort.

Usage: DATABASE_URL=... python load_gsa_roll_xlsx.py <fips> <roll.zip|.xlsx> [--dry-run]
"""
import io
import json
import os
import re
import sys
import time
import zipfile
from datetime import date, datetime

import openpyxl
import psycopg2

POOL_RE = re.compile(r"pool|swim", re.I)
GARAGE_RE = re.compile(r"garage|carport", re.I)
SHED_RE = re.compile(r"\bshed\b|workshop|out ?building|storage|\bbarn\b|stable|boat *dock", re.I)
SKIP_DESC = {"", "RESIDENCE", "MOBILE HOME", "MULTI FAM RES", "MAIN AREA"}


def to_int(v, lo, hi):
    try:
        n = int(float(str(v).strip()))
    except (ValueError, TypeError):
        return None
    return n if lo <= n <= hi else None


def parse_dt(v):
    if isinstance(v, (datetime, date)):
        d = v.date() if isinstance(v, datetime) else v
        return d if 1900 <= d.year <= date.today().year else None
    s = str(v or "").strip()
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m/%d/%y"):
        try:
            d = datetime.strptime(s.split()[0], fmt).date()
            return d if 1900 <= d.year <= date.today().year else None
        except (ValueError, IndexError):
            continue
    return None


def open_sheet(path):
    if path.lower().endswith(".zip"):
        z = zipfile.ZipFile(path)
        member = next(n for n in z.namelist() if n.lower().endswith(".xlsx"))
        data = io.BytesIO(z.read(member))
        wb = openpyxl.load_workbook(data, read_only=True)
    else:
        wb = openpyxl.load_workbook(path, read_only=True)
    return wb[wb.sheetnames[0]]


def main():
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    dry = "--dry-run" in sys.argv
    fips, src = args[0], args[1]
    ws = open_sheet(src)
    it = ws.iter_rows(values_only=True)
    hdr = [str(h).strip() if h is not None else "" for h in next(it)]
    ix = {h: i for i, h in enumerate(hdr)}
    ACC, GEO = ix.get("ACCOUNT_NUMBER"), ix.get("GIS_IDENTIFICATION_NUM")
    DS, EX = ix.get("DATE_OF_SALE"), ix.get("EXEMPTION")
    IT, IA, YB = ix.get("IMPROVEMENT_TYPE"), ix.get("IMPROVEMENT_AREA"), ix.get("YEAR_BUILT")

    staged = {}   # account -> dict
    t0 = time.time()
    n = 0
    for r in it:
        if ACC is None or len(r) <= ACC or not r[ACC]:
            continue
        n += 1
        acc = str(r[ACC]).strip()
        geo = str(r[GEO]).strip() if GEO is not None and r[GEO] else ""
        e = staged.get(acc)
        if e is None:
            e = staged[acc] = {"geo": geo, "descs": set(), "sqft": 0, "yr": None,
                               "dt": None, "ex": set()}
        elif geo and not e["geo"]:
            e["geo"] = geo
        desc = str(r[IT]).strip() if IT is not None and r[IT] else ""
        if desc and desc.upper() not in SKIP_DESC:
            e["descs"].add(desc)
        a = to_int(r[IA], 1, 2_000_000) if IA is not None else None
        if a:
            e["sqft"] += a
        y = to_int(r[YB], 1800, date.today().year + 1) if YB is not None else None
        if y:
            e["yr"] = max(e["yr"] or 0, y)
        d = parse_dt(r[DS]) if DS is not None else None
        if d and (e["dt"] is None or d > e["dt"]):
            e["dt"] = d
        exv = str(r[EX]).strip().upper() if EX is not None and r[EX] else ""
        for tok in re.split(r"[,;/\s]+", exv):
            tok = tok.strip()
            if tok and tok not in ("0", "NONE"):
                e["ex"].add(tok)
    print(f"[{fips}] parsed {n:,} rows -> {len(staged):,} accounts ({time.time()-t0:.0f}s)", flush=True)
    _db_phase(fips, staged, dry)


def _connect():
    for a in range(6):
        try:
            c = psycopg2.connect(os.environ["DATABASE_URL"], keepalives=1, keepalives_idle=30,
                                 keepalives_interval=10, keepalives_count=10, connect_timeout=20)
            cur = c.cursor(); cur.execute("SET lock_timeout='5s'"); c.commit()
            return c
        except psycopg2.OperationalError as e:
            print(f"  connect retry {a+1}: {e}", flush=True); time.sleep(5)
    raise RuntimeError("could not connect")


def _db_phase(fips, staged, dry):
    conn = _connect(); cur = conn.cursor()
    cur.execute("""CREATE TEMP TABLE gsa_stage (acc TEXT PRIMARY KEY, geo TEXT,
                       improvements JSONB, sqft INT, yr INT, sale_dt DATE,
                       exemptions TEXT[], pool BOOL, garage BOOL, shed BOOL)""")
    buf = io.StringIO()
    for acc, e in staged.items():
        descs = sorted(e["descs"]) if e["descs"] else None
        blob = " ".join(descs).lower() if descs else ""
        arr = "{" + ",".join(sorted(e["ex"])) + "}" if e["ex"] else r"\N"
        buf.write("\t".join([
            acc.replace("\\", "\\\\"),
            e["geo"].replace("\\", "\\\\") if e["geo"] else r"\N",
            json.dumps(descs).replace("\\", "\\\\") if descs else r"\N",
            str(e["sqft"]) if e["sqft"] else r"\N",
            r"\N" if e["yr"] is None else str(e["yr"]),
            r"\N" if e["dt"] is None else e["dt"].isoformat(),
            arr,
            "t" if POOL_RE.search(blob) else "f",
            "t" if GARAGE_RE.search(blob) else "f",
            "t" if SHED_RE.search(blob) else "f",
        ]) + "\n")
    buf.seek(0)
    cur.copy_expert("COPY gsa_stage (acc, geo, improvements, sqft, yr, sale_dt, exemptions, pool, garage, shed) FROM STDIN", buf)
    cur.execute("ANALYZE gsa_stage")

    cands = [("spid==acct", "p.source_property_id = s.acc"),
             ("apn==acct", "p.apn = s.acc"),
             ("spid==geoid", "p.source_property_id = s.geo"),
             ("apn==geoid", "p.apn = s.geo")]
    best_join, best_n = None, 0
    for label, cond in cands:
        cur.execute(f"SELECT count(*) FROM gsa_stage s JOIN parcels p ON p.county_fips=%s AND {cond}", (fips,))
        nj = cur.fetchone()[0]
        print(f"  join {label}: {nj:,}", flush=True)
        if nj > best_n:
            best_n, best_join = nj, cond
    cur.execute("SELECT count(*) FROM parcels WHERE county_fips=%s", (fips,))
    tot = cur.fetchone()[0]
    if best_n < 0.30 * max(tot, 1):
        conn.rollback(); conn.close()
        raise SystemExit(f"[{fips}] ABORT: best join {best_n:,} < 30% of {tot:,}")
    print(f"  -> {best_join} ({best_n:,}/{tot:,})", flush=True)
    if dry:
        conn.rollback(); conn.close(); print("dry-run"); return

    cur.execute(f"""UPDATE parcels p SET
                        improvements     = COALESCE(p.improvements, s.improvements),
                        living_area_sqft = COALESCE(p.living_area_sqft, s.sqft),
                        year_built       = COALESCE(p.year_built, s.yr),
                        last_sale_date   = COALESCE(p.last_sale_date, s.sale_dt),
                        exemptions       = COALESCE(s.exemptions, p.exemptions),
                        has_pool         = COALESCE(p.has_pool, FALSE) OR s.pool,
                        has_garage       = COALESCE(p.has_garage, FALSE) OR s.garage,
                        has_shed         = COALESCE(p.has_shed, FALSE) OR s.shed
                    FROM gsa_stage s WHERE p.county_fips=%s AND {best_join}""", (fips,))
    print(f"  parcels updated: {cur.rowcount:,}", flush=True)
    conn.commit()
    cur.execute("""SELECT count(*) FILTER (WHERE improvements IS NOT NULL), count(*) FILTER (WHERE living_area_sqft IS NOT NULL),
                          count(*) FILTER (WHERE last_sale_date IS NOT NULL), count(*) FILTER (WHERE array_length(exemptions,1)>0)
                   FROM parcels WHERE county_fips=%s""", (fips,))
    imp, sq, sd, ex = cur.fetchone()
    print(f"[{fips}] now: improv {imp:,}, sqft {sq:,}, sale {sd:,}, exempt {ex:,}", flush=True)
    conn.close()


if __name__ == "__main__":
    main()
